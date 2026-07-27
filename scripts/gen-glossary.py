"""從 docs/glossary-terms-tracker.xlsx 產生 src/data/glossary.ts。

欄位對應：
  A 狀態(不進網站)  B 類型→kind        C 縮寫／用詞→abbr   D 英文全名→en
  E 日文名稱→ja     F 中文名稱→zh      G 簡易說明→description
  H 別名／搜尋字→aliases              I 網站已收錄(不進網站)
  J 來源路徑→sources                 K 備註→family（只取「XX家族」）
日文以試算表 E 欄為準；該欄留空時，退回從 type-content.ts 的 ssbu-05 名詞清單比對帶入。
"""
import openpyxl, re, json, collections

ROOT = '/Users/yuhudaddy/Desktop/yuda website'
GAMES = {'BotW': 'botw', 'TotK': 'totk', 'EoW': 'eow', 'SSBU': 'ssbu', 'AoC': 'aoc', 'AoI': 'aoi'}

# 試算表所有有「縮寫／用詞」的列都會上架，「狀態」欄僅供作者自己追蹤進度。
# 中文名稱留空（或填「-」）的條目仍會輸出，頁面上該欄留白。

KIND = {
    '遊戲名稱縮寫': 'game',
    '技巧縮寫': 'abbr',
    '原理機制': 'concept',
    '物件／狀態': 'object',
}

FAMILY = {
    '過載': ('overload', '過載'),
    '硬直取消': ('endlag', '硬直取消'),
    '隱藏': ('cull', '隱藏'),
    '穿牆': ('clip', '穿牆'),
    '餘料': ('fuse', '餘料'),
    '糾纏': ('fuse', '餘料'),
    '擊飛': ('launch', '擊飛'),
    '複製': ('dupe', '複製'),
    '並列': ('zuggle', '並列'),
    '黏手': ('zuggle', '並列'),
    '突擊': ('flurry', '突擊'),
    '災禍': ('curse', '災禍'),
    '懲戒': ('purgatory', '懲戒'),
    'SBR': ('sbr', '盾擋重置'),
}

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return None if s in ('', '-', '—') else s

# ── 1. 讀 ssbu-05 的日文 ──
# 站內把多個相關詞併成一條（例：「Hitbox / Hurtbox / Disjoint」對「攻撃判定 / やられ判定 / 武器判定」），
# 兩邊同樣以「/」分隔且順序一致，故可逐項拆開對應回單一術語。
JA_ALIAS = {'Risk and Reward': 'Risk-Reward', '? + tilt/smash/air': '招式方向縮寫'}

src = open(f'{ROOT}/src/data/type-content.ts', encoding='utf-8').read()
seg = src[src.index('"ssbu-05"'):src.index('"ssbu-06"')]
ja_map = {}
for zh, en, ja in re.findall(r'\{ zh: "([^"]*)"(?:, en: "([^"]*)")?(?:, ja: "([^"]*)")?', seg):
    if not ja:
        continue
    zh, ja = zh.strip(), ja.strip()
    zh = JA_ALIAS.get(zh, zh)
    zh_parts = [p.strip() for p in zh.split(' / ')]
    ja_parts = [p.strip() for p in ja.split(' / ')]
    if len(ja_parts) >= len(zh_parts) > 1:
        for zp, jp in zip(zh_parts, ja_parts):
            ja_map[zp] = jp
    else:
        for zp in zh_parts:
            ja_map[zp] = ja
    ja_map.setdefault(zh, ja)

# ── 2. 讀 xlsx ──
wb = openpyxl.load_workbook(f'{ROOT}/docs/glossary-terms-tracker.xlsx')
merged = {}
order = []
skipped = 0
no_zh = []
for sheet, game in GAMES.items():
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        c = [clean(ws.cell(row=r, column=i).value) for i in range(1, 12)]
        state, kind_zh, abbr, en, ja_cell, zh, desc, aliases, _, path, note = c
        if not abbr:
            skipped += 1
            continue
        if not zh:
            no_zh.append(f'{sheet}!{abbr}')
        if path == '/types/glossary':
            path = None       # 指向術語對照頁本身，不需要列成來源連結

        fam = None
        m = re.search(r'([一-鿿A-Za-z]+)\s*家族', note or '')
        if m:
            fam = FAMILY.get(m.group(1))

        key = (abbr, zh)
        if key in merged:                       # 跨遊戲同義詞合併
            e = merged[key]
            if game not in e['games']:
                e['games'].append(game)
            if path and path not in e['sources']:
                e['sources'].append(path)
            if en and (not e.get('en') or len(en) > len(e['en'])):
                e['en'] = en
            if aliases:
                for a in (x.strip() for x in aliases.split(',')):
                    if a and a not in e['aliases']:
                        e['aliases'].append(a)
            e.setdefault('family', fam and fam[0])
            continue

        e = {
            'games': [game],
            'kind': KIND.get(kind_zh, 'term'),
            'abbr': abbr,
            'en': en,
            # 以試算表的日文欄為準；留空時才回頭查站內名詞表。
            # 後備查詢限定大亂鬥，否則同名術語會誤植
            # （例：DI 在王淚是 Despawn Interrupt，與大亂鬥的「ベク変」無關）
            'ja': ja_cell or (ja_map.get(abbr) if game == 'ssbu' else None),
            'zh': zh,
            'description': desc,
            'aliases': [x.strip() for x in aliases.split(',')] if aliases else [],
            'family': fam[0] if fam else None,
            'sources': [path] if path else [],
        }
        merged[key] = e
        order.append(key)

entries = [merged[k] for k in order]

# ── 3. 產生唯一 id ──
seen = collections.Counter()
for e in entries:
    base = re.sub(r'[^a-z0-9]+', '-', e['abbr'].lower()).strip('-') or e['games'][0]
    seen[base] += 1
    e['id'] = base if seen[base] == 1 else f"{base}-{e['games'][0]}"

# ── 4. 輸出 ──
def ts(v):
    return json.dumps(v, ensure_ascii=False)

lines = []
w = lines.append
w('// 由 docs/glossary-terms-tracker.xlsx 產生，請勿手改；')
w('// 要新增或修改術語請改試算表後重新產生。')
w('')
w('export type GlossaryGameId = "botw" | "totk" | "eow" | "ssbu" | "aoc" | "aoi";')
w('')
w('export interface GlossaryGame {')
w('  id: GlossaryGameId;')
w('  tag: string;')
w('  label: string;')
w('  title: string;')
w('  aliases: string[];')
w('}')
w('')
w('export interface GlossaryEntry {')
w('  id: string;')
w('  games: GlossaryGameId[];')
w('  kind: "game" | "abbr" | "concept" | "object" | "term";')
w('  abbr: string;        // 索引主鍵：縮寫、原文或中文詞')
w('  en?: string;         // 英文全名（站內未提供者省略）')
w('  ja?: string;         // 日文（目前僅大亂鬥術語有）')
w('  zh?: string;         // 中文名稱（試算表尚未填寫者省略）')
w('  aliases?: string[];')
w('  family?: string;     // 技巧家族 id，對應 glossaryFamilies')
w('  description?: string;')
w('  sources?: string[];')
w('}')
w('')
w('// 技巧家族：站內多數術語成組出現，索引頁用來分組與交叉篩選')
w('export const glossaryFamilies: Record<string, string> = {')
for fid, flabel in sorted({v for v in FAMILY.values()}):
    w(f'  {fid}: {ts(flabel)},')
w('};')
w('')

GAME_META = [
    ('botw', 'BotW', '曠野之息', 'Breath of the Wild', ['BoTW', 'Zelda BotW', '薩爾達傳說 曠野之息']),
    ('totk', 'TotK', '王國之淚', 'Tears of the Kingdom', ['ToTK', 'Zelda TotK', '薩爾達傳說 王國之淚']),
    ('eow', 'EoW', '智慧的再現', 'Echoes of Wisdom', ['Zelda EoW', '薩爾達傳說 智慧的再現']),
    ('ssbu', 'SSBU', '任天堂明星大亂鬥 特別版', 'Super Smash Bros. Ultimate', ['SSB Ultimate', '大亂鬥 SP', 'Smash Ultimate']),
    ('aoc', 'AoC', '災厄啟示錄', 'Age of Calamity', ['Zelda AoC', 'Hyrule Warriors: Age of Calamity']),
    ('aoi', 'AoI', '封印戰記', 'Age of Imprisonment', ['Zelda AoI', 'Hyrule Warriors: Age of Imprisonment']),
]
w('export const glossaryGames: GlossaryGame[] = [')
for gid, tag, label, title, al in GAME_META:
    w(f'  {{ id: "{gid}", tag: {ts(tag)}, label: {ts(label)}, title: {ts(title)}, aliases: {ts(al)} }},')
w('];')
w('')
w('export const glossaryEntries: GlossaryEntry[] = [')
for e in entries:
    parts = [f'id: {ts(e["id"])}', f'games: {ts(e["games"])}', f'kind: {ts(e["kind"])}', f'abbr: {ts(e["abbr"])}']
    if e['en']:
        parts.append(f'en: {ts(e["en"])}')
    if e['ja']:
        parts.append(f'ja: {ts(e["ja"])}')
    if e['zh']:
        parts.append(f'zh: {ts(e["zh"])}')
    if e['aliases']:
        parts.append(f'aliases: {ts(e["aliases"])}')
    if e['family']:
        parts.append(f'family: {ts(e["family"])}')
    if e['description']:
        parts.append(f'description: {ts(e["description"])}')
    if e['sources']:
        parts.append(f'sources: {ts(e["sources"])}')
    w('  { ' + ', '.join(parts) + ' },')
w('];')
w('')

open(f'{ROOT}/src/data/glossary.ts', 'w', encoding='utf-8').write('\n'.join(lines))

# ── 統計 ──
print(f'條目：{len(entries)}（試算表另有 {skipped} 列無縮寫／用詞而略過）')
print(f'有英文全名：{sum(1 for e in entries if e["en"])}')
print(f'有日文：    {sum(1 for e in entries if e["ja"])}')
print(f'有說明：    {sum(1 for e in entries if e["description"])}')
print(f'有家族：    {sum(1 for e in entries if e["family"])}')
print('kind 分布：', dict(collections.Counter(e['kind'] for e in entries)))
print('多遊戲條目：', [f'{e["abbr"]}({"+".join(e["games"])})' for e in entries if len(e['games']) > 1])
if no_zh:
    print(f'\n⚠ 尚無中文名稱（頁面該欄留白）：{len(no_zh)} 筆')
    for x in no_zh: print('   ', x)
